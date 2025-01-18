import random
import csv
from datetime import datetime
from openpyxl import Workbook

# Define languages and their codes
LANGUAGES = {
    'en': 'English',
    'ar': 'Arabic',
    'fr': 'French'
}

# Experience templates for each category
ACADEMIC_SUCCESS = {
    'en': [
        "After months of dedicated preparation and countless late-night study sessions, I achieved a perfect score on my thesis defense. The committee was particularly impressed with my innovative research methodology and the practical applications of my findings. This success has opened up several opportunities for publication in prestigious journals.",
        "My research project on sustainable energy solutions won the university's annual innovation award. The project involved developing a new solar cell technology that improved efficiency by 25%. This recognition has led to funding opportunities and collaboration offers from industry partners.",
        "Successfully completed my semester with a 4.0 GPA while managing three advanced research projects. The key was developing an effective time management system and building strong relationships with my professors. This achievement has boosted my confidence in pursuing even more challenging academic goals."
    ],
    'ar': [
        "بعد شهور من التحضير المكثف وجلسات الدراسة الليلية، حققت درجة كاملة في دفاع أطروحتي. أبدت اللجنة إعجابها بشكل خاص بمنهجية البحث المبتكرة والتطبيقات العملية لنتائجي. هذا النجاح فتح العديد من الفرص للنشر في المجلات المرموقة.",
        "فاز مشروع بحثي حول حلول الطاقة المستدامة بجائزة الابتكار السنوية للجامعة. تضمن المشروع تطوير تقنية خلايا شمسية جديدة حسنت الكفاءة بنسبة 25%. أدى هذا الاعتراف إلى فرص تمويل وعروض تعاون من شركاء الصناعة.",
        "أكملت الفصل الدراسي بنجاح بمعدل 4.0 مع إدارة ثلاثة مشاريع بحثية متقدمة. كان المفتاح هو تطوير نظام فعال لإدارة الوقت وبناء علاقات قوية مع أساتذتي. عزز هذا الإنجاز ثقتي في متابعة أهداف أكاديمية أكثر تحديًا."
    ],
    'fr': [
        "Après des mois de préparation intensive et d'innombrables sessions d'étude nocturnes, j'ai obtenu une note parfaite à ma soutenance de thèse. Le comité a été particulièrement impressionné par ma méthodologie de recherche innovante et les applications pratiques de mes résultats. Ce succès m'a ouvert plusieurs opportunités de publication dans des revues prestigieuses.",
        "Mon projet de recherche sur les solutions énergétiques durables a remporté le prix annuel de l'innovation de l'université. Le projet impliquait le développement d'une nouvelle technologie de cellules solaires qui améliorait l'efficacité de 25%. Cette reconnaissance a conduit à des opportunités de financement et des offres de collaboration avec des partenaires industriels.",
        "J'ai terminé mon semestre avec une moyenne de 4.0 tout en gérant trois projets de recherche avancés. La clé était de développer un système efficace de gestion du temps et de construire des relations solides avec mes professeurs. Cette réussite a renforcé ma confiance pour poursuivre des objectifs académiques encore plus ambitieux."
    ]
}

RESEARCH_EXPERIENCES = {
    'en': [
        "Leading a groundbreaking research project in quantum computing has been both challenging and rewarding. Our team recently made a breakthrough in quantum error correction that could significantly impact the field. We're now preparing to present our findings at an international conference, and I'm learning so much about project management and team leadership.",
        "My involvement in a cross-disciplinary research initiative combining biology and artificial intelligence has opened new horizons. We're developing novel algorithms for protein folding prediction, and our preliminary results are very promising. The experience of working with experts from different fields has enriched my understanding of collaborative research.",
        "Currently working on an environmental research project studying climate change impacts on local ecosystems. We've collected data over two years, and our findings are revealing concerning patterns. This experience has taught me the importance of long-term data collection and careful methodology in scientific research."
    ],
    'ar': [
        "قيادة مشروع بحثي رائد في مجال الحوسبة الكمية كان تحديًا ومجزيًا في آن واحد. حقق فريقنا مؤخرًا اختراقًا في تصحيح الأخطاء الكمية يمكن أن يؤثر بشكل كبير على المجال. نحن نستعد الآن لتقديم نتائجنا في مؤتمر دولي، وأتعلم الكثير عن إدارة المشاريع وقيادة الفريق.",
        "مشاركتي في مبادرة بحثية متعددة التخصصات تجمع بين علم الأحياء والذكاء الاصطناعي فتحت آفاقًا جديدة. نطور خوارزميات جديدة للتنبؤ بطي البروتين، ونتائجنا الأولية واعدة جدًا. أثرت تجربة العمل مع خبراء من مجالات مختلفة فهمي للبحث التعاوني.",
        "أعمل حاليًا على مشروع بحثي بيئي يدرس آثار تغير المناخ على النظم البيئية المحلية. جمعنا البيانات على مدى عامين، وتكشف نتائجنا عن أنماط مثيرة للقلق. علمتني هذه التجربة أهمية جمع البيانات على المدى الطويل والمنهجية الدقيقة في البحث العلمي."
    ],
    'fr': [
        "Diriger un projet de recherche novateur en informatique quantique a été à la fois stimulant et gratifiant. Notre équipe a récemment fait une percée dans la correction d'erreurs quantiques qui pourrait avoir un impact significatif sur le domaine. Nous nous préparons maintenant à présenter nos résultats lors d'une conférence internationale, et j'apprends beaucoup sur la gestion de projet et le leadership d'équipe.",
        "Ma participation à une initiative de recherche interdisciplinaire combinant biologie et intelligence artificielle a ouvert de nouveaux horizons. Nous développons de nouveaux algorithmes pour la prédiction du repliement des protéines, et nos résultats préliminaires sont très prometteurs. L'expérience de travailler avec des experts de différents domaines a enrichi ma compréhension de la recherche collaborative.",
        "Je travaille actuellement sur un projet de recherche environnementale étudiant les impacts du changement climatique sur les écosystèmes locaux. Nous avons collecté des données sur deux ans, et nos résultats révèlent des tendances préoccupantes. Cette expérience m'a appris l'importance de la collecte de données à long terme et d'une méthodologie rigoureuse dans la recherche scientifique."
    ]
}

CULTURAL_EXPERIENCES = {
    'en': [
        "Being part of the international student community has transformed my worldview completely. I've participated in cultural exchange programs, learned to cook dishes from five different countries, and made friends from across the globe. These experiences have taught me the value of diversity and cross-cultural communication in ways I never expected.",
        "Organizing our university's first multicultural festival was an incredible experience. We showcased traditions from over 20 countries, including traditional dances, music performances, and food tastings. The event brought together hundreds of students and helped bridge cultural gaps on campus.",
        "My semester abroad in a different country pushed me far outside my comfort zone. From navigating daily life in a new language to adapting to different academic expectations, every day brought new challenges and learning opportunities. This experience has made me more adaptable and culturally aware."
    ],
    'ar': [
        "كوني جزءًا من مجتمع الطلاب الدوليين غير نظرتي للعالم تمامًا. شاركت في برامج التبادل الثقافي، وتعلمت طهي أطباق من خمس دول مختلفة، وكونت صداقات من جميع أنحاء العالم. علمتني هذه التجارب قيمة التنوع والتواصل بين الثقافات بطرق لم أتوقعها أبدًا.",
        "كان تنظيم أول مهرجان متعدد الثقافات في جامعتنا تجربة لا تصدق. عرضنا تقاليد من أكثر من 20 دولة، بما في ذلك الرقصات التقليدية والعروض الموسيقية وتذوق الطعام. جمع الحدث مئات الطلاب وساعد في سد الفجوات الثقافية في الحرم الجامعي.",
        "دفعني فصلي الدراسي في الخارج بعيدًا عن منطقة الراحة الخاصة بي. من التنقل في الحياة اليومية بلغة جديدة إلى التكيف مع توقعات أكاديمية مختلفة، كل يوم جلب تحديات وفرص تعلم جديدة. جعلتني هذه التجربة أكثر قدرة على التكيف ووعيًا ثقافيًا."
    ],
    'fr': [
        "Faire partie de la communauté des étudiants internationaux a complètement transformé ma vision du monde. J'ai participé à des programmes d'échange culturel, appris à cuisiner des plats de cinq pays différents et me suis fait des amis du monde entier. Ces expériences m'ont appris la valeur de la diversité et de la communication interculturelle d'une manière que je n'aurais jamais imaginée.",
        "L'organisation du premier festival multiculturel de notre université a été une expérience incroyable. Nous avons présenté des traditions de plus de 20 pays, incluant des danses traditionnelles, des performances musicales et des dégustations culinaires. L'événement a rassemblé des centaines d'étudiants et a aidé à combler les écarts culturels sur le campus.",
        "Mon semestre à l'étranger m'a poussé bien au-delà de ma zone de confort. De la navigation dans la vie quotidienne dans une nouvelle langue à l'adaptation à différentes attentes académiques, chaque jour apportait de nouveaux défis et opportunités d'apprentissage. Cette expérience m'a rendu plus adaptable et culturellement conscient."
    ]
}

WORK_LIFE_BALANCE = {
    'en': [
        "Managing a part-time job at the university library while maintaining my academic performance has been a valuable learning experience. I've developed excellent time management skills and learned to prioritize tasks effectively. The experience has also taught me the importance of setting boundaries and taking care of my mental health.",
        "Balancing my role as a teaching assistant with my graduate studies has been challenging but rewarding. I've learned to structure my days efficiently, dedicating specific time blocks for research, teaching preparation, and personal development. This experience has helped me become more organized and resilient.",
        "Finding harmony between my academic commitments and personal life took time to master. I've developed a system where I schedule regular exercise sessions, social activities, and study periods. This balanced approach has actually improved both my academic performance and overall well-being."
    ],
    'ar': [
        "كانت إدارة وظيفة بدوام جزئي في مكتبة الجامعة مع الحفاظ على أدائي الأكاديمي تجربة تعليمية قيمة. طورت مهارات ممتازة في إدارة الوقت وتعلمت تحديد أولويات المهام بشكل فعال. علمتني التجربة أيضًا أهمية وضع الحدود والاهتمام بصحتي النفسية.",
        "كان التوازن بين دوري كمساعد تدريس ودراساتي العليا تحديًا ولكنه مجزٍ. تعلمت تنظيم أيامي بكفاءة، مخصصًا فترات زمنية محددة للبحث والتحضير للتدريس والتطوير الشخصي. ساعدتني هذه التجربة على أن أصبح أكثر تنظيمًا ومرونة.",
        "استغرق العثور على التناغم بين التزاماتي الأكاديمية وحياتي الشخصية وقتًا لإتقانه. طورت نظامًا أجدول فيه جلسات التمرين المنتظمة والأنشطة الاجتماعية وفترات الدراسة. حسن هذا النهج المتوازن من أدائي الأكاديمي ورفاهيتي العامة."
    ],
    'fr': [
        "Gérer un emploi à temps partiel à la bibliothèque universitaire tout en maintenant mes performances académiques a été une expérience d'apprentissage précieuse. J'ai développé d'excellentes compétences en gestion du temps et appris à prioriser efficacement les tâches. L'expérience m'a également appris l'importance de fixer des limites et de prendre soin de ma santé mentale.",
        "Équilibrer mon rôle d'assistant d'enseignement avec mes études supérieures a été difficile mais gratifiant. J'ai appris à structurer mes journées efficacement, en consacrant des blocs de temps spécifiques à la recherche, à la préparation de l'enseignement et au développement personnel. Cette expérience m'a aidé à devenir plus organisé et résilient.",
        "Trouver l'harmonie entre mes engagements académiques et ma vie personnelle a pris du temps à maîtriser. J'ai développé un système où je planifie des sessions d'exercice régulières, des activités sociales et des périodes d'étude. Cette approche équilibrée a en fait amélioré à la fois mes performances académiques et mon bien-être général."
    ]
}

PERSONAL_GROWTH = {
    'en': [
        "This year of graduate studies has been transformative for my personal development. Through various challenges and successes, I've learned to embrace uncertainty and view obstacles as opportunities for growth. The support from my mentors and peers has been instrumental in building my confidence and resilience.",
        "Taking on leadership roles in student organizations has helped me discover strengths I didn't know I had. From organizing events to mediating conflicts, each experience has contributed to my growth as a leader. I've learned that effective leadership is about empowering others and fostering collaboration.",
        "My journey of self-discovery in academia has been profound. I've overcome my fear of public speaking through regular presentations, developed critical thinking skills through research projects, and learned to advocate for myself and others. These experiences have shaped not just my academic career, but my entire perspective on life."
    ],
    'ar': [
        "كان هذا العام من الدراسات العليا تحويليًا لتطوري الشخصي. من خلال التحديات والنجاحات المختلفة، تعلمت تقبل عدم اليقين والنظر إلى العقبات كفرص للنمو. كان دعم موجهيّ وزملائي أساسيًا في بناء ثقتي ومرونتي.",
        "ساعدني تولي أدوار قيادية في المنظمات الطلابية على اكتشاف نقاط قوة لم أكن أعرف أنني أمتلكها. من تنظيم الفعاليات إلى حل النزاعات، ساهمت كل تجربة في نموي كقائد. تعلمت أن القيادة الفعالة تتعلق بتمكين الآخرين وتعزيز التعاون.",
        "كانت رحلة اكتشاف الذات في الأوساط الأكاديمية عميقة. تغلبت على خوفي من التحدث أمام الجمهور من خلال العروض التقديمية المنتظمة، وطورت مهارات التفكير النقدي من خلال المشاريع البحثية، وتعلمت الدفاع عن نفسي والآخرين. شكلت هذه التجارب ليس فقط مسيرتي الأكاديمية، ولكن منظوري الكامل للحياة."
    ],
    'fr': [
        "Cette année d'études supérieures a été transformatrice pour mon développement personnel. À travers divers défis et succès, j'ai appris à embrasser l'incertitude et à voir les obstacles comme des opportunités de croissance. Le soutien de mes mentors et pairs a été déterminant dans le renforcement de ma confiance et de ma résilience.",
        "Assumer des rôles de leadership dans les organisations étudiantes m'a aidé à découvrir des forces que je ne savais pas avoir. De l'organisation d'événements à la médiation de conflits, chaque expérience a contribué à ma croissance en tant que leader. J'ai appris que le leadership efficace consiste à autonomiser les autres et à favoriser la collaboration.",
        "Mon parcours de découverte de soi dans le milieu universitaire a été profond. J'ai surmonté ma peur de parler en public grâce à des présentations régulières, développé des compétences de pensée critique à travers des projets de recherche, et appris à défendre mes intérêts et ceux des autres. Ces expériences ont façonné non seulement ma carrière académique, mais aussi toute ma perspective sur la vie."
    ]
}

def generate_experiences():
    experiences = []
    categories = {
        'Academic Success': ACADEMIC_SUCCESS,
        'Research Experiences': RESEARCH_EXPERIENCES,
        'Cultural Experiences': CULTURAL_EXPERIENCES,
        'Work-Life Balance': WORK_LIFE_BALANCE,
        'Personal Growth': PERSONAL_GROWTH
    }
    
    # Generate 12 experiences from each category
    for category, templates in categories.items():
        for _ in range(12):
            lang = random.choice(list(LANGUAGES.keys()))
            text = random.choice(templates.get(lang, templates['en']))
            experiences.append({
                'text': text,
                'category': category,
                'language': LANGUAGES[lang]
            })
    
    # Shuffle the experiences
    random.shuffle(experiences)
    return experiences

def save_to_csv(experiences, base_filename):
    filename = f'{base_filename}.csv'
    with open(filename, 'w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=['text', 'category', 'language'])
        writer.writeheader()
        writer.writerows(experiences)
    return filename

def save_to_txt(experiences, base_filename):
    filename = f'{base_filename}.txt'
    with open(filename, 'w', encoding='utf-8') as file:
        for exp in experiences:
            file.write(f"{exp['text']}\n")
    return filename

def save_to_excel(experiences, base_filename):
    filename = f'{base_filename}.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Experiences"
    
    # Write headers
    headers = ['Text', 'Category', 'Language']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row, exp in enumerate(experiences, 2):
        ws.cell(row=row, column=1, value=exp['text'])
        ws.cell(row=row, column=2, value=exp['category'])
        ws.cell(row=row, column=3, value=exp['language'])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(filename)
    return filename

if __name__ == '__main__':
    experiences = generate_experiences()
    base_filename = f'testdata'
    
    csv_file = save_to_csv(experiences, base_filename)
    txt_file = save_to_txt(experiences, base_filename)
    excel_file = save_to_excel(experiences, base_filename)
    
    print(f"Generated {len(experiences)} experiences and saved to:")
    print(f"CSV: {csv_file}")
    print(f"TXT: {txt_file}")
    print(f"Excel: {excel_file}")
