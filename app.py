import streamlit as st
import pandas as pd
import google.generativeai as genai
import os
import time
from pathlib import Path
import io
from streamlit_extras.switch_page_button import switch_page
import openpyxl

# Constants
MIN_BATCH_SIZE = 10
MAX_BATCH_SIZE = 20

# Configure Gemini API and page settings
st.set_page_config(
    page_title="تجارب الطلاب",
    page_icon="🎓",
    layout="centered"
)

# Add custom CSS for RTL support and centering
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700&display=swap');

    /* Global RTL settings */
    .main {
        direction: rtl !important;
        font-family: 'Cairo', sans-serif !important;
        padding: 1em !important;
        font-size: 16px !important;
    }
    
    /* Text alignment for all elements */
    .stMarkdown, .stText, .stTitle, div[data-testid="stText"], .stProgress > div > div {
        text-align: right !important;
        direction: rtl !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 1.1em !important;
    }

    /* Enhanced RTL for all streamlit elements */
    .element-container, .stTextInput, .stSelectbox, .stDateInput {
        direction: rtl !important;
        text-align: right !important;
    }

    /* Increase font size for all inputs */
    .stTextInput input, .stSelectbox select, .stDateInput input {
        font-size: 1.1em !important;
        font-family: 'Cairo', sans-serif !important;
    }

    /* RTL for file uploader */
    .stFileUploader {
        direction: rtl !important;
        text-align: right !important;
        font-size: 1.1em !important;
    }

    /* RTL for radio buttons */
    .stRadio > div {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for dataframe */
    .dataframe {
        direction: rtl !important;
        text-align: right !important;
        font-size: 1.1em !important;
    }

    /* Increase font size for buttons */
    .stButton>button {
        font-size: 1.2em !important;
        padding: 1em 1.5em !important;
    }

    /* Headers with increased font size */
    h1 {
        font-size: 2.8em !important;
    }

    h2 {
        font-size: 2.3em !important;
    }

    h3 {
        font-size: 1.8em !important;
    }

    /* Experience cards with increased font size */
    .experience-card {
        font-size: 1.1em !important;
    }

    /* Summary section with increased font size */
    .summary-container {
        font-size: 1.1em !important;
    }

    .stat-number {
        font-size: 2.8em !important;
    }

    .stat-label {
        font-size: 1.4em !important;
    }

    /* Category cards with increased font size */
    .category-name {
        font-size: 1.1em !important;
    }

    .category-count {
        font-size: 1.8em !important;
    }

    /* Toast messages with increased font size */
    .toast {
        font-size: 1.2em !important;
        text-align: right !important;
        direction: rtl !important;
    }

    /* Preview section with increased font size */
    .preview-header {
        font-size: 1.6em !important;
    }

    .preview-content {
        font-size: 1.1em !important;
    }

    /* Processing text with increased font size */
    .processing-text {
        font-size: 1.4em !important;
        text-align: right !important;
    }

    /* RTL for select boxes */
    div[data-baseweb="select"] {
        direction: rtl !important;
    }

    div[data-baseweb="select"] > div {
        text-align: right !important;
    }

    /* RTL for tooltips */
    .stTooltipIcon {
        direction: rtl !important;
    }

    /* RTL for tabs */
    .stTabs [data-baseweb="tab-list"] {
        direction: rtl !important;
    }

    /* RTL for metrics */
    .stMetric {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for expander */
    .streamlit-expanderHeader {
        direction: rtl !important;
        text-align: right !important;
        font-size: 1.2em !important;
    }

    /* Button alignment and styling */
    .stButton>button {
        float: right !important;
        width: 100% !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 1.1em !important;
        padding: 0.8em 1.5em !important;
        background-color: #2196F3 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 5px rgba(33, 150, 243, 0.3) !important;
    }

    .stButton>button:hover {
        background-color: #1976D2 !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 4px 8px rgba(33, 150, 243, 0.4) !important;
    }
    
    /* Header and toolbar RTL */
    div[data-testid="stToolbar"], div[data-testid="stHeader"] {
        direction: rtl !important;
    }
    
    /* Headings styling */
    h1, h2, h3 {
        font-family: 'Cairo', sans-serif !important;
        font-weight: 700 !important;
        color: #1f1f1f !important;
        text-align: right !important;
        margin-bottom: 0.8em !important;
        padding-right: 0.5em !important;
        border-right: 4px solid #2196F3 !important;
    }

    h1 {
        font-size: 2.5em !important;
    }

    h2 {
        font-size: 2em !important;
    }
    
    /* File uploader styling */
    .stFileUploader {
        padding: 1.5em !important;
        background: #f8f9fa !important;
        border-radius: 12px !important;
        border: 2px dashed #dee2e6 !important;
        margin: 1em 0 !important;
    }

    .stFileUploader:hover {
        border-color: #2196F3 !important;
        background: #f1f8fe !important;
    }

    /* Select box styling */
    .stSelectbox > div > div {
        font-family: 'Cairo', sans-serif !important;
        border-radius: 8px !important;
    }
    
    /* Radio button styling */
    div.row-widget.stRadio > div {
        display: flex !important;
        flex-direction: row !important;
        justify-content: flex-start !important;
        gap: 2em !important;
        direction: rtl !important;
        margin: 1em 0 !important;
    }
    
    div.row-widget.stRadio > div[role="radiogroup"] > label {
        background: #f8f9fa !important;
        padding: 0.7em 1.2em !important;
        border-radius: 8px !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        text-align: center !important;
        font-family: 'Cairo', sans-serif !important;
        border: 1px solid #e0e0e0 !important;
        flex: 1 !important;
        min-width: 120px !important;
        direction: rtl !important;
    }
    
    div.row-widget.stRadio > div[role="radiogroup"] > label:hover {
        background: #e9ecef !important;
        transform: translateY(-2px) !important;
        border-color: #2196F3 !important;
    }
    
    div.row-widget.stRadio > div[role="radiogroup"] > label[data-baseweb="radio"] > div:first-child {
        display: none !important;
    }

    /* Radio button label styling */
    .stRadio > label {
        text-align: right !important;
        width: 100% !important;
        display: block !important;
        margin-bottom: 0.5em !important;
        font-family: 'Cairo', sans-serif !important;
        font-weight: 600 !important;
        color: #1f1f1f !important;
        direction: rtl !important;
    }

    /* Preview section styling */
    .preview-section {
        margin: 2em 0;
        background: #ffffff;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(33, 150, 243, 0.08);
        border: 1px solid #e0e0e0;
        overflow: hidden;
    }

    .preview-header-container {
        background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%);
        padding: 1em 1.5em;
        border-bottom: 1px solid #e0e0e0;
        display: flex;
        align-items: center;
        justify-content: flex-start;
        direction: rtl;
    }

    .preview-icon {
        font-size: 1.5em;
        margin-left: 0.8em;
        background: linear-gradient(45deg, #2196F3, #64B5F6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        animation: float 3s ease-in-out infinite;
    }

    @keyframes float {
        0% { transform: translateY(0px); }
        50% { transform: translateY(-5px); }
        100% { transform: translateY(0px); }
    }

    .preview-title {
        font-family: 'Cairo', sans-serif;
        font-size: 1.4em;
        font-weight: 600;
        color: #1f1f1f;
        margin: 0;
        position: relative;
    }

    .preview-title::after {
        content: '';
        position: absolute;
        bottom: -5px;
        right: 0;
        width: 40px;
        height: 3px;
        background: linear-gradient(90deg, #2196F3, #64B5F6);
        border-radius: 2px;
        transition: width 0.3s ease;
    }

    .preview-header-container:hover .preview-title::after {
        width: 100%;
    }

    .preview-content {
        padding: 1.5em;
        background: #ffffff;
    }

    @media (max-width: 768px) {
        .preview-header-container {
            padding: 1em;
        }
        
        .preview-title {
            font-size: 1.2em;
        }
        
        .preview-icon {
            font-size: 1.3em;
        }
    }

    /* Form elements RTL */
    .stSelectbox, .stTextInput {
        direction: rtl !important;
        text-align: right !important;
        font-family: 'Cairo', sans-serif !important;
    }

    /* Progress bar styling */
    .stProgress > div > div {
        background-color: #2196F3 !important;
        height: 8px !important;
        border-radius: 4px !important;
    }

    .stProgress > div {
        background-color: #e3f2fd !important;
        border-radius: 4px !important;
    }

    /* Experience cards styling */
    .experience-card {
        background: #ffffff;
        border-radius: 12px;
        padding: 1.5em;
        margin: 1em 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border: 1px solid #e0e0e0;
        font-family: 'Cairo', sans-serif !important;
        transition: all 0.3s ease;
    }

    .experience-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }

    .experience-card.positive {
        border-right: 5px solid #4CAF50;
    }

    .experience-card.negative {
        border-right: 5px solid #f44336;
    }

    /* Statistics box styling */
    .stats-box {
        background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%);
        border-radius: 12px;
        padding: 1.5em;
        margin-bottom: 1em !important;
        text-align: center;
        font-family: 'Cairo', sans-serif !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        border: 1px solid #e0e0e0;
    }

    .stats-number {
        font-size: 2.5em;
        font-weight: bold;
        margin: 0.5em 0;
        color: #2196F3;
        text-shadow: 1px 1px 0 rgba(0,0,0,0.1);
    }

    /* Form label styling */
    .stSelectbox > label, 
    .stFileUploader > label,
    div[data-baseweb="select"] > label,
    .stRadio > label {
        font-size: 1.4em !important;
        font-weight: 600 !important;
        color: #1f1f1f !important;
        margin-bottom: 0.8em !important;
        display: block !important;
        font-family: 'Cairo', sans-serif !important;
    }

    /* Progress bar container */
    .stProgress {
        margin: 2em 0 !important;
    }

    /* Progress bar styling */
    .stProgress > div > div {
        background: linear-gradient(90deg, #2196F3 0%, #64B5F6 50%, #2196F3 100%) !important;
        background-size: 200% 100% !important;
        animation: progress-pulse 2s ease-in-out infinite !important;
        height: 10px !important;
        border-radius: 5px !important;
    }

    @keyframes progress-pulse {
        0% { background-position: 100% 0; }
        100% { background-position: -100% 0; }
    }

    .stProgress > div {
        background-color: #e3f2fd !important;
        border-radius: 5px !important;
        height: 10px !important;
    }

    /* Progress message */
    .progress-message {
        text-align: center !important;
        margin: 1em 0 !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 1.1em !important;
        color: #1976D2 !important;
    }

    /* Toast styling */
    .toast {
        position: fixed !important;
        top: 80px !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
        padding: 1em 1.5em !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 1.1em !important;
        margin-bottom: 10px !important;
        z-index: 999999 !important;
        min-width: 300px !important;
        max-width: 80% !important;
        animation: toast-slide 3s forwards !important;
    }

    @keyframes toast-slide {
        0% { transform: translate(-50%, -100%); opacity: 0; }
        10% { transform: translate(-50%, 0); opacity: 1; }
        90% { transform: translate(-50%, 0); opacity: 1; }
        100% { transform: translate(-50%, -100%); opacity: 0; }
    }

    .toast.success {
        background-color: #4CAF50 !important;
        color: white !important;
        border-right: 4px solid #2E7D32 !important;
    }

    .toast.error {
        background-color: #f44336 !important;
        color: white !important;
        border-right: 4px solid #C62828 !important;
    }

    /* Tabs styling for RTL */
    .stTabs {
        direction: rtl !important;
    }

    .stTabs > div[role="tablist"] {
        direction: rtl !important;
        justify-content: flex-start !important;
        gap: 1em !important;
    }

    .stTabs [role="tab"] {
        font-family: 'Cairo', sans-serif !important;
        font-size: 1.1em !important;
        padding: 0.5em 1em !important;
        border-radius: 8px !important;
    }

    /* Info bubble styling */
    .experience-card {
        position: relative !important;
    }

    .info-link-container {
        position: absolute !important;
        left: 15px !important;
        top: 15px !important;
    }

    .info-link {
        display: inline-block !important;
        width: 24px !important;
        height: 24px !important;
        background: #2196F3 !important;
        color: white !important;
        border-radius: 50% !important;
        text-align: center !important;
        line-height: 24px !important;
        text-decoration: none !important;
        font-weight: bold !important;
    }

    .info-bubble {
        display: none !important;
        position: absolute !important;
        left: 30px !important;
        top: 0 !important;
        background: white !important;
        border: 1px solid #e0e0e0 !important;
        border-radius: 8px !important;
        padding: 10px !important;
        width: 250px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        z-index: 1000 !important;
    }

    .info-link:hover + .info-bubble {
        display: block !important;
    }

    .spinner {
        width: 50px;
        height: 50px;
        border: 5px solid #f3f3f3;
        border-top: 5px solid #2196F3;
        border-radius: 50%;
        animation: spin 1s linear infinite;
        margin: 20px auto;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    .processing-text {
        text-align: center;
        color: #2196F3;
        font-family: 'Cairo', sans-serif;
        margin: 10px 0;
        font-size: 1.2em;
    }

    /* Remove toast styling */
    .stSuccess {
        font-family: 'Cairo', sans-serif !important;
        direction: rtl !important;
        text-align: right !important;
        padding: 1em !important;
        border-radius: 8px !important;
        margin: 1em 0 !important;
    }

    /* Summary section styling */
    .summary-container {
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
        border-radius: 15px;
        padding: 30px;
        margin: 20px 0;
        box-shadow: 0 4px 15px rgba(0,0,0,0.05);
        border: 1px solid #e0e0e0;
    }

    .summary-header {
        text-align: center;
        color: #1f1f1f;
        font-size: 1.8em;
        font-weight: 700;
        margin-bottom: 30px;
        padding-bottom: 15px;
        border-bottom: 2px solid #e0e0e0;
    }

    .summary-stats {
        display: flex;
        justify-content: center;
        gap: 40px;
        margin-bottom: 30px;
        padding: 20px;
        background: #ffffff;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    .stat-card {
        text-align: center;
        padding: 15px 25px;
        border-radius: 10px;
        transition: transform 0.3s ease;
    }

    .stat-card:hover {
        transform: translateY(-5px);
    }

    .stat-number {
        font-size: 2.5em;
        font-weight: 700;
        margin-bottom: 10px;
    }

    .positive-stat .stat-number {
        color: #4CAF50;
        text-shadow: 0 2px 4px rgba(76, 175, 80, 0.2);
    }

    .negative-stat .stat-number {
        color: #f44336;
        text-shadow: 0 2px 4px rgba(244, 67, 54, 0.2);
    }

    .stat-label {
        font-size: 1.2em;
        color: #666;
        font-weight: 600;
    }

    .top-categories {
        margin-top: 25px;
        background: #ffffff;
        border-radius: 12px;
        padding: 15px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    .top-categories-header {
        text-align: center;
        font-size: 1.4em;
        font-weight: 600;
        margin-bottom: 20px;
        color: #1f1f1f;
        position: relative;
    }

    .category-grid {
        display: flex;
        justify-content: space-between;
        align-items: stretch;
        gap: 10px;
        padding: 0 5px;
    }

    .category-card {
        flex: 1;
        width: calc(25% - 8px); /* 25% width for 4 cards with gap consideration */
        min-width: 120px;
        max-width: 180px;
        background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%);
        padding: 12px 10px;
        border-radius: 8px;
        text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        border: 1px solid #e0e0e0;
        transition: all 0.3s ease;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
    }

    .category-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(33, 150, 243, 0.15);
        border-color: #2196F3;
    }

    .category-name {
        font-weight: 600;
        color: #2196F3;
        margin-top: 5px;
        font-size: 0.9em;
        line-height: 1.3;
        word-wrap: break-word;
        max-width: 100%;
    }

    .category-count {
        font-size: 1.6em;
        color: #1f1f1f;
        font-weight: 700;
        line-height: 1;
        background: linear-gradient(135deg, #2196F3, #64B5F6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 2px;
    }

    @media (max-width: 768px) {
        .category-grid {
            gap: 8px;
        }
        
        .category-card {
            padding: 10px 6px;
            min-width: 90px;
        }
        
        .category-name {
            font-size: 0.85em;
        }
        
        .category-count {
            font-size: 1.4em;
        }
    }

    /* Download button styling */
    .stDownloadButton {
        width: 100%;
        margin: 20px 0;
    }

    .stDownloadButton button {
        width: 100% !important;
        font-family: 'Cairo', sans-serif !important;
        background: linear-gradient(135deg, #2196F3 0%, #1976D2 100%) !important;
        color: white !important;
        padding: 15px 30px !important;
        border: none !important;
        border-radius: 10px !important;
        font-size: 1.2em !important;
        font-weight: 600 !important;
        cursor: pointer !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 15px rgba(33, 150, 243, 0.3) !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        gap: 10px !important;
    }

    .stDownloadButton button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(33, 150, 243, 0.4) !important;
        background: linear-gradient(135deg, #1976D2 0%, #1565C0 100%) !important;
    }

    .stDownloadButton button::before {
        content: "📥" !important;
        font-size: 1.2em !important;
    }

    /* Category percentages styling */
    .category-percentages {
        margin-top: 5px;
        font-size: 0.8em;
        display: flex;
        flex-direction: column;
        gap: 2px;
        text-align: center;
    }

    .positive-pct {
        color: #4CAF50;
    }

    .negative-pct {
        color: #f44336;
    }

    .category-card {
        padding: 15px 10px !important;
    }

    /* Radio button and select box labels with larger font */
    div.row-widget.stRadio > div[role="radiogroup"] > label {
        font-size: 1.3em !important;
        padding: 0.8em 1.5em !important;
    }

    .stSelectbox > label {
        font-size: 1.5em !important;
        font-weight: 600 !important;
        margin-bottom: 1em !important;
        color: #1f1f1f !important;
    }

    /* Radio group label */
    .stRadio > label {
        font-size: 1.5em !important;
        font-weight: 600 !important;
        margin-bottom: 1em !important;
        color: #1f1f1f !important;
    }

    /* Processing section styling */
    .processing-container {
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        margin: 20px auto !important;
        width: 100% !important;
        max-width: 500px !important;
        position: relative !important;
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%) !important;
        border-radius: 15px !important;
        padding: 30px !important;
        box-shadow: 0 4px 15px rgba(33, 150, 243, 0.1) !important;
        overflow: hidden !important;
    }

    .processing-container::before {
        content: '' !important;
        position: absolute !important;
        top: 0 !important;
        left: -100% !important;
        width: 200% !important;
        height: 100% !important;
        background: linear-gradient(
            90deg,
            transparent,
            rgba(33, 150, 243, 0.1),
            transparent
        ) !important;
        animation: wave 2s infinite !important;
    }

    @keyframes wave {
        0% { transform: translateX(0); }
        100% { transform: translateX(50%); }
    }

    .loader {
        width: 80px !important;
        height: 80px !important;
        position: relative !important;
        margin: 20px auto !important;
    }

    .loader-ring {
        width: 100% !important;
        height: 100% !important;
        border-radius: 50% !important;
        position: absolute !important;
        top: 0 !important;
        left: 0 !important;
    }

    .loader-ring-1 {
        border: 3px solid rgba(33, 150, 243, 0.1) !important;
        border-top: 3px solid #2196F3 !important;
        animation: spin 1s linear infinite !important;
    }

    .loader-ring-2 {
        border: 3px solid transparent !important;
        border-top: 3px solid #64B5F6 !important;
        animation: spin 1s linear infinite reverse !important;
        width: 70% !important;
        height: 70% !important;
        margin: 15% !important;
    }

    .loader-ring-3 {
        border: 3px solid rgba(33, 150, 243, 0.1) !important;
        border-top: 3px solid #2196F3 !important;
        animation: spin 1.5s linear infinite !important;
        width: 40% !important;
        height: 40% !important;
        margin: 30% !important;
    }

    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }

    .processing-text {
        text-align: center !important;
        color: #2196F3 !important;
        font-family: 'Cairo', sans-serif !important;
        font-size: 1.4em !important;
        font-weight: 600 !important;
        margin: 15px 0 !important;
        background: linear-gradient(45deg, #2196F3, #64B5F6) !important;
        -webkit-background-clip: text !important;
        -webkit-text-fill-color: transparent !important;
        animation: pulse 2s infinite !important;
    }

    @keyframes pulse {
        0% { opacity: 0.6; }
        50% { opacity: 1; }
        100% { opacity: 0.6; }
    }

    .processing-progress {
        width: 80% !important;
        height: 6px !important;
        background: rgba(33, 150, 243, 0.1) !important;
        border-radius: 3px !important;
        margin: 15px auto !important;
        overflow: hidden !important;
        position: relative !important;
    }

    .progress-bar {
        position: absolute !important;
        width: 50% !important;
        height: 100% !important;
        background: linear-gradient(90deg, #2196F3, #64B5F6) !important;
        border-radius: 3px !important;
        animation: progress 2s ease-in-out infinite !important;
    }

    @keyframes progress {
        0% { left: -50%; }
        100% { left: 100%; }
    }

    /* Preview section styling */
    .preview-content {
        max-height: 300px;
        overflow-y: auto;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        padding: 10px;
        margin: 10px 0;
        background: #ffffff;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }

    /* Enhanced text preview styling */
    .txt-preview {
        display: block;
        direction: rtl;
        position: relative;
        padding: 0;
        background: #ffffff;
        font-family: 'Cairo', sans-serif;
        line-height: 1.6;
    }

    .responses-container {
        padding: 15px 45px 15px 15px;
        counter-reset: response-counter;
    }

    .response-item {
        position: relative;
        padding: 8px 15px;
        margin: 5px 0;
        border-radius: 6px;
        background: #f8f9fa;
        transition: all 0.2s ease;
        border-right: 3px solid #2196F3;
        counter-increment: response-counter;
    }

    .response-item::before {
        content: counter(response-counter);
        position: absolute;
        right: -30px;
        top: 50%;
        transform: translateY(-50%);
        color: #666;
        font-size: 0.9em;
        width: 20px;
        text-align: left;
    }

    .response-item:hover {
        background: #f1f8fe;
        transform: translateX(-2px);
        box-shadow: 0 2px 5px rgba(33, 150, 243, 0.1);
    }

    /* Scrollbar styling for the preview */
    .txt-preview::-webkit-scrollbar {
        width: 8px;
    }

    .txt-preview::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 4px;
    }

    .txt-preview::-webkit-scrollbar-thumb {
        background: #2196F3;
        border-radius: 4px;
    }

    .txt-preview::-webkit-scrollbar-thumb:hover {
        background: #1976D2;
    }

    /* Ensure text wrapping */
    .response-item {
        white-space: pre-wrap;
        word-wrap: break-word;
    }

    @keyframes pulseBox {
        0% { box-shadow: 0 0 0 0 rgba(33, 150, 243, 0.4); }
        70% { box-shadow: 0 0 0 10px rgba(33, 150, 243, 0); }
        100% { box-shadow: 0 0 0 0 rgba(33, 150, 243, 0); }
    }
    
    .important-note {
        background: linear-gradient(135deg, #e3f2fd 0%, #ffffff 100%);
        padding: 20px;
        border-radius: 12px;
        border-right: 4px solid #2196F3;
        margin: 20px 0;
        box-shadow: 0 4px 15px rgba(33, 150, 243, 0.1);
        animation: pulseBox 2s infinite;
        position: relative;
        overflow: hidden;
    }
    
    .important-note::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 2px;
        background: linear-gradient(90deg, #2196F3, #64B5F6);
    }
    
    .note-icon {
        font-size: 1.5em;
        margin-left: 10px;
        color: #2196F3;
        vertical-align: middle;
    }
    
    .note-header {
        font-family: 'Cairo', sans-serif;
        font-size: 1.3em;
        font-weight: 700;
        color: #1565C0;
        margin-bottom: 8px;
        display: flex;
        align-items: center;
    }
    
    .note-content {
        font-family: 'Cairo', sans-serif;
        font-size: 1.1em;
        color: #1f1f1f;
        line-height: 1.6;
        margin: 0;
        padding-right: 15px;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(-10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .important-note {
        animation: fadeIn 0.5s ease-out, pulseBox 2s infinite;
    }

    /* Header styling */
    .header-container {
        background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%);
        padding: 2em;
        border-radius: 15px;
        margin-bottom: 2em;
        box-shadow: 0 4px 15px rgba(33, 150, 243, 0.1);
        border: 1px solid #e0e0e0;
        position: relative;
        overflow: hidden;
    }
    
    .header-container::before {
        content: '';
        position: absolute;
        top: 0;
        right: 0;
        width: 6px;
        height: 100%;
        background: linear-gradient(180deg, #2196F3, #64B5F6);
        border-radius: 3px;
    }
    
    .header-icon {
        font-size: 2.5em;
        margin-bottom: 0.3em;
        background: linear-gradient(45deg, #2196F3, #64B5F6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        display: inline-block;
    }
    
    .header-title {
        font-family: 'Cairo', sans-serif;
        font-size: 2.5em;
        font-weight: 700;
        color: #1f1f1f;
        margin: 0;
        padding: 0;
        line-height: 1.2;
    }
    
    .header-subtitle {
        font-family: 'Cairo', sans-serif;
        font-size: 1.2em;
        color: #666;
        margin-top: 0.5em;
        line-height: 1.4;
    }
    
    .header-content {
        text-align: right;
        direction: rtl;
    }
    
    @media (max-width: 768px) {
        .header-container {
            padding: 1.5em;
        }
        
        .header-title {
            font-size: 2em;
        }
        
        .header-subtitle {
            font-size: 1em;
        }
    }
</style>
""", unsafe_allow_html=True)


#------------------------------------------------------------------------------
# Gemini Communication
#------------------------------------------------------------------------------

def upload_to_gemini(path, mime_type=None):
    """Uploads the given file to Gemini."""
    try:
        # Use pathlib for cross-platform path handling
        base_path = Path(__file__).parent
        file_path = base_path / path
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        file = genai.upload_file(str(file_path), mime_type=mime_type)
        print(f"Uploaded file '{file.display_name}' as: {file.uri}")
        return file
    except Exception as e:
        st.error(f"Failed to upload file: {e}")
        return None

def wait_for_files_active(files):
    """Waits for the given files to be active."""
    print("Waiting for file processing...")
    try:
        for file in files:
            if file is None:
                raise Exception("Invalid file object")

            name = file.name
            file = genai.get_file(name)
            while file.state.name == "PROCESSING":
                print(".", end="", flush=True)
                time.sleep(2)
                file = genai.get_file(name)
            if file.state.name != "ACTIVE":
                raise Exception(f"File {file.name} failed to process")
        print("...all files ready")
        print()
    except Exception as e:
        st.error(f"File processing failed: {e}")
        return False
    return True

def initialize_gemini():
    """Initialize Gemini model with API key from streamlit secrets."""
    try:
        # Get API key from streamlit secrets
        api_key = st.secrets["GEMINI_API_KEY"]
        if not api_key:
            st.error("API key not found in secrets. Please check your .streamlit/secrets.toml file.")
            return None

        genai.configure(api_key=api_key)

        # Create the model
        generation_config = {
            "temperature": 0,
            "top_p": 0.95,
            "top_k": 40,
            "max_output_tokens": 8192,
            "response_mime_type": "application/json",
        }

        model = genai.GenerativeModel(
            model_name="gemini-1.5-flash",
            generation_config=generation_config,
            system_instruction=(
                "according to the categories mentinoed. which category does the provided text fit in the most? "
                "what is the most appropriate subcategory? and what is the most appropriate type (positive or negative)? "
                "you must use a category, subcategory, and type from the file only, choose from them what fits the case the most. "
                "the output should be in arabic. make the a json object. "
                "the keys are: category, subcategory, type, explanation. "
            )
        )

        # Upload and process the categories file
        files = [
            upload_to_gemini("data/Classes.txt", mime_type="text/plain"),
        ]

        # Check if file upload was successful
        if None in files:
            raise Exception("Failed to upload required files")

        # Wait for files to be processed
        if not wait_for_files_active(files):
            raise Exception("File processing failed")

        chat_session = model.start_chat(
            history=[
                {
                    "role": "user",
                    "parts": [
                        files[0],
                    ],
                },
            ]
        )
        return chat_session
    except Exception as e:
        st.error(f"Failed to initialize Gemini: {e}")
        return None

def read_csv_with_encoding(file):
    """Read CSV file with multiple encoding attempts."""
    encodings = ['utf-8', 'utf-8-sig', 'cp1256', 'iso-8859-6']
    df = None
    
    for encoding in encodings:
        try:
            df = pd.read_csv(file, encoding=encoding)
            if not df.empty and len(df.columns) > 0:
                return df
        except Exception:
            continue
    
    return None

def process_responses(file, file_type, column_name=None, separator=None):
    """Process uploaded file and extract responses."""
    try:
        # Clear preview data from other tabs
        st.session_state.preview_data = None
        
        if file_type == 'txt':
            content = file.getvalue().decode('utf-8')
            responses = [r.strip() for r in content.split(separator) if r.strip()]
        elif file_type == 'csv':
            if st.session_state.current_df is None:
                raise Exception("لم يتم تحميل الملف بشكل صحيح")
            
            df = st.session_state.current_df
            responses = df[column_name].dropna().tolist()
        elif file_type == 'excel':
            df = pd.read_excel(file)
            if df.empty or len(df.columns) == 0:
                raise Exception("لم يتم العثور على أعمدة صالحة في الملف. تأكد من تنسيق ملف Excel")
            responses = df[column_name].dropna().tolist()
        
        # Update preview data
        preview_df = pd.DataFrame({"الاستجابات": responses})
        st.session_state.preview_data = preview_df
        
        return responses
    except pd.errors.EmptyDataError:
        st.error("الملف فارغ أو لا يحتوي على بيانات صالحة")
        return []
    except KeyError:
        st.error(f"لم يتم العثور على العمود '{column_name}' في الملف")
        return []
    except Exception as e:
        st.error(f"حدث خطأ أثناء معالجة الملف: {str(e)}")
        return []

def get_optimal_batch_size(total_items):
    """Calculate optimal batch size based on total items."""
    # For very small datasets
    if total_items <= MIN_BATCH_SIZE:
        return total_items
    
    # Calculate a reasonable batch size (about 1/3 of total)
    suggested_size = max(MIN_BATCH_SIZE, total_items // 3)
    
    # Cap at MAX_BATCH_SIZE
    return min(suggested_size, MAX_BATCH_SIZE)

def classify_responses_batch(responses_batch):
    """Classify a batch of responses using Gemini."""
    try:
        if not st.session_state.model:
            raise Exception("نموذج Gemini غير مهيأ")
        
        # Combine responses into a single request with clear separation
        batch_text = "\n=====\n".join([f"Response {i+1}: {str(r)}" for i, r in enumerate(responses_batch)])
        
        # Create prompt with escaped curly braces for the example
        prompt = (
            f"Please classify these {len(responses_batch)} responses. For each response, determine:\n"
            "1. Category (التصنيف)\n"
            "2. Subcategory (التصنيف_فرعي)\n"
            "3. Type (نوع) - must be either 'إيجابي' or 'سلبي'\n"
            "4. Explanation (تفسير)\n\n"
            "Return a JSON array with one object per response. Example format:\n"
            "[\n"
            "    {{\n"
            '        "category": "example_category",\n'
            '        "subcategory": "example_subcategory",\n'
            '        "type": "إيجابي",\n'
            '        "explanation": "example_explanation"\n'
            "    }}\n"
            "]\n\n"
            "Here are the responses to classify:\n\n"
            f"{batch_text}"
        )

        # Send request to model
        chat_response = st.session_state.model.send_message(prompt)
        
        try:
            # Try to parse as JSON first
            import json
            classifications = json.loads(chat_response.text)
            
        except json.JSONDecodeError as e:
            # Try to clean the response and parse again
            try:
                # Remove any markdown formatting if present
                cleaned_text = chat_response.text.strip('`').strip()
                if cleaned_text.startswith('json'):
                    cleaned_text = cleaned_text[4:].strip()
                classifications = json.loads(cleaned_text)
            except json.JSONDecodeError as e2:
                st.markdown("""
                    <div class="toast error">
                        فشل في تحليل استجابة النموذج: تنسيق JSON غير صالح
                    </div>
                """, unsafe_allow_html=True)
                return [None] * len(responses_batch)
        
        # Ensure we have a list
        if not isinstance(classifications, list):
            st.markdown("""
                <div class="toast error">
                    استجابة النموذج ليست قائمة صالحة
                </div>
            """, unsafe_allow_html=True)
            classifications = [classifications]
        
        # Validate and normalize each classification
        validated_classifications = []
        for i, classification in enumerate(classifications):
            try:
                if not isinstance(classification, dict):
                    validated_classifications.append(None)
                    continue
                
                # Get type value, handling different possible keys
                type_value = classification.get('type', classification.get('نوع', ''))
                
                # Normalize type values
                if type_value.strip() in ['إيجابي', 'سلبي', 'ايجابية', 'سلبية', 'ايجابي', 'إيجابية']:
                    normalized_type = 'إيجابي' if type_value.strip() in ['إيجابي', 'ايجابية', 'ايجابي'] else 'سلبي'
                    
                    validated_classifications.append({
                        'type': normalized_type,
                        'category': classification.get('category', classification.get('تصنيف', '')).strip(),
                        'subcategory': classification.get('subcategory', classification.get('تصنيف_فرعي', '')).strip(),
                        'explanation': classification.get('explanation', classification.get('تفسير', '')).strip()
                    })
                else:
                    validated_classifications.append(None)
            except Exception as e:
                validated_classifications.append(None)
        
        return validated_classifications
        
    except Exception as e:
        st.markdown(f"""
            <div class="toast error">
                خطأ في التصنيف: {str(e)}
            </div>
        """, unsafe_allow_html=True)
        return [None] * len(responses_batch)

def display_experience(result, experience_type):
    """Enhanced display function for experiences"""
    try:
        card_class = "positive" if experience_type == "positive" else "negative"
        
        st.markdown(f"""
            <div class="experience-card {card_class}">
                <div class="response-text">
                    <strong>الاستجابة:</strong> {result.get('response', '')}
                </div>
                <div class="category-text">
                    <strong>التصنيف:</strong> {result.get('classification', {}).get('category', '')}
                </div>
                <div class="category-text">
                    <strong>التصنيف الفرعي:</strong> {result.get('classification', {}).get('subcategory', '')}
                </div>
                <div class="info-link-container">
                    <a href="#" class="info-link">
                        <span class="info-icon">i</span>
                    </a>
                    <div class="info-bubble">
                        {result.get('classification', {}).get('explanation', 'لا يوجد تفسير')}
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)
    except Exception as e:
        st.error(f"خطأ في عرض التجربة: {str(e)}")

# Initialize session state variables
if 'preview_data' not in st.session_state:
    st.session_state.preview_data = None
if 'model' not in st.session_state:
    st.session_state.model = None
if 'classification_results' not in st.session_state:
    st.session_state.classification_results = None
if 'current_df' not in st.session_state:
    st.session_state.current_df = None
if 'results' not in st.session_state:
    st.session_state.results = []
if 'previous_file_type' not in st.session_state:
    st.session_state.previous_file_type = None

# Initialize Gemini at startup if not already initialized
if st.session_state.model is None:
    st.session_state.model = initialize_gemini()

st.markdown("""
    <div class="header-container">
        <div class="header-content">
            <div class="header-icon">🎓</div>
            <h1 class="header-title">محلل تجارب الطلاب</h1>
            <div class="header-subtitle">تحليل وتصنيف تجارب الطلاب باستخدام الذكاء الاصطناعي</div>
        </div>
    </div>
""", unsafe_allow_html=True)

# Replace tabs with select box for file type
file_type = st.selectbox(
    "اختر نوع الملف",
    ["ملف CSV", "ملف نصي", "ملف Excel"],
    key="file_type_selector"
)

# Clear preview data when file type changes
if st.session_state.previous_file_type != file_type:
    st.session_state.preview_data = None
    st.session_state.current_df = None
    st.session_state.previous_file_type = file_type

responses = []
if file_type == "ملف نصي":
    st.markdown("""
        <style>
            @keyframes pulseBox {
                0% { box-shadow: 0 0 0 0 rgba(33, 150, 243, 0.4); }
                70% { box-shadow: 0 0 0 10px rgba(33, 150, 243, 0); }
                100% { box-shadow: 0 0 0 0 rgba(33, 150, 243, 0); }
            }
            
            .important-note {
                background: linear-gradient(135deg, #e3f2fd 0%, #ffffff 100%);
                padding: 20px;
                border-radius: 12px;
                border-right: 4px solid #2196F3;
                margin: 20px 0;
                box-shadow: 0 4px 15px rgba(33, 150, 243, 0.1);
                animation: pulseBox 2s infinite;
                position: relative;
                overflow: hidden;
            }
            
            .important-note::before {
                content: '';
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                height: 2px;
                background: linear-gradient(90deg, #2196F3, #64B5F6);
            }
            
            .note-icon {
                font-size: 1.5em;
                margin-left: 10px;
                color: #2196F3;
                vertical-align: middle;
            }
            
            .note-header {
                font-family: 'Cairo', sans-serif;
                font-size: 1.3em;
                font-weight: 700;
                color: #1565C0;
                margin-bottom: 8px;
                display: flex;
                align-items: center;
            }
            
            .note-content {
                font-family: 'Cairo', sans-serif;
                font-size: 1.1em;
                color: #1f1f1f;
                line-height: 1.6;
                margin: 0;
                padding-right: 15px;
            }
            
            @keyframes fadeIn {
                from { opacity: 0; transform: translateY(-10px); }
                to { opacity: 1; transform: translateY(0); }
            }
            
            .important-note {
                animation: fadeIn 0.5s ease-out, pulseBox 2s infinite;
            }
        </style>
        <div class="important-note">
            <div class="note-header">
                <span class="note-icon">ℹ️</span>
                <span>تنبيه مهم</span>
            </div>
            <p class="note-content">
                يجب أن تكون كل استجابة في سطر جديد في الملف النصي للمعالجة الصحيحة للبيانات
            </p>
        </div>
    """, unsafe_allow_html=True)
    
    txt_file = st.file_uploader("تحميل ملف نصي", type=['txt'], key="txt_uploader")
    if txt_file:
        responses = process_responses(txt_file, 'txt', separator="\n")

elif file_type == "ملف CSV":
    csv_file = st.file_uploader("تحميل ملف CSV", type=['csv'], key="csv_uploader")
    if csv_file:
        df = read_csv_with_encoding(csv_file)
        if df is not None:
            st.session_state.current_df = df
            columns = df.columns.tolist()
            column_name = st.selectbox("حدد العمود الذي يحتوي على استجابات الطلاب:", columns, key="csv_column")
            if column_name:
                responses = process_responses(csv_file, 'csv', column_name=column_name)
        else:
            st.error("لم يتم العثور على أعمدة صالحة في الملف. تأكد من تنسيق الملف CSV وترميزه")

else:  # Excel file
    excel_file = st.file_uploader("تحميل ملف Excel", type=['xlsx'], key="excel_uploader")
    if excel_file:
        df = pd.read_excel(excel_file)
        columns = df.columns.tolist()
        column_name = st.selectbox("حدد العمود الذي يحتوي على استجابات الطلاب:", columns, key="excel_column")
        if column_name:
            responses = process_responses(excel_file, 'excel', column_name=column_name)

# Preview section with compact design
if st.session_state.preview_data is not None:
    st.markdown("""
        <style>
        .preview-content {
            max-height: 300px;
            overflow-y: auto;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            padding: 10px;
            margin: 10px 0;
        }
        
        .compact-preview {
            height: 100%;
        }
        
        .compact-table {
            width: 100%;
            border-collapse: collapse;
        }
        
        .compact-table th {
            background: #f8f9fa;
            position: sticky;
            top: 0;
            z-index: 1;
            padding: 10px;
            border-bottom: 2px solid #dee2e6;
        }
        
        .compact-table td {
            padding: 8px 10px;
            border-bottom: 1px solid #e9ecef;
        }
        
        .preview-text {
            white-space: pre-wrap;
            font-family: 'Cairo', sans-serif;
            line-height: 1.6;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Wrap preview in collapsible section
    st.markdown("""
        <div class="preview-section">
            <div class="preview-header-container">
                <div class="preview-icon">📊</div>
                <div class="preview-title">معاينة البيانات</div>
            </div>
    """, unsafe_allow_html=True)
    
    if file_type == "ملف نصي":
        # Create numbered list of responses
        responses_html = ""
        for i, response in enumerate(st.session_state.preview_data["الاستجابات"], 1):
            responses_html += f'<div class="response-item">{response}</div>'
            
        st.markdown(f"""
            <div class="txt-preview">
                <div class="responses-container">
                    {responses_html}
                </div>
            </div>
        """, unsafe_allow_html=True)
    else:
        st.dataframe(
            st.session_state.preview_data,
            use_container_width=True,
            height=300,
            hide_index=True
        )
    
    st.markdown('</div></div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        if st.button("تصنيف الاستجابات", key="classify_button"):
            if not st.session_state.model:
                st.markdown("""
                    <div class="toast error">
                        لم يتم تهيئة نموذج Gemini بشكل صحيح. يرجى التحقق من مفتاح API
                    </div>
                """, unsafe_allow_html=True)
            else:
                processing_container = st.empty()
                processing_container.markdown("""
                    <div class="processing-container">
                        <div class="loader">
                            <div class="loader-ring loader-ring-1"></div>
                            <div class="loader-ring loader-ring-2"></div>
                            <div class="loader-ring loader-ring-3"></div>
                        </div>
                        <div class="processing-text">جاري معالجة الاستجابات...</div>
                        <div class="processing-progress">
                            <div class="progress-bar"></div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
                
                results = []
                
                try:
                    # Calculate optimal batch size
                    batch_size = get_optimal_batch_size(len(responses))
                    total_batches = (len(responses) + batch_size - 1) // batch_size
                    
                    # Process responses in batches
                    for i in range(0, len(responses), batch_size):
                        batch = responses[i:i + batch_size]
                        classifications = classify_responses_batch(batch)
                        
                        for response, classification in zip(batch, classifications):
                            if classification:
                                results.append({
                                    "response": response,
                                    "classification": classification
                                })
                    
                    # Clear processing indicator
                    processing_container.empty()
                    
                    # Store results in session state for persistence
                    st.session_state.classification_results = results
                    st.session_state.results = results
                    
                    # Show success message
                    st.markdown("""
                        <div class="toast success">
                            تم تصنيف الاستجابات بنجاح!
                        </div>
                    """, unsafe_allow_html=True)

                except Exception as e:
                    processing_container.empty()
                    st.markdown(f"""
                        <div class="toast error">
                            حدث خطأ أثناء التصنيف: {str(e)}
                        </div>
                    """, unsafe_allow_html=True)

# After classification is complete, display results
if st.session_state.get('results'):
    st.markdown('<h2 class="centered">نتائج التصنيف</h2>', unsafe_allow_html=True)
    
    # Filter experiences
    positive_experiences = [r for r in st.session_state.results if r.get('classification', {}).get('type') == 'إيجابي']
    negative_experiences = [r for r in st.session_state.results if r.get('classification', {}).get('type') == 'سلبي']
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create a combined DataFrame with all results
        all_results_df = pd.DataFrame([
            {
                'الاستجابة': r['response'],
                'التصنيف': r['classification']['category'],
                'التصنيف الفرعي': r['classification']['subcategory'],
                'النوع': r['classification']['type'],
                'التفسير': r['classification']['explanation']
            }
            for r in st.session_state.results
        ])

        # Add timestamp
        all_results_df['تاريخ التحليل'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')

        # Create summary sheet
        summary_sheet = pd.DataFrame()
        total_responses = len(all_results_df)
        positive_count = len(all_results_df[all_results_df['النوع'] == 'إيجابي'])
        negative_count = len(all_results_df[all_results_df['النوع'] == 'سلبي'])
        
        summary_data = {
            'المقياس': ['إجمالي الاستجابات', 'التجارب الإيجابية', 'التجارب السلبية', 
                      'نسبة التجارب الإيجابية', 'نسبة التجارب السلبية'],
            'القيمة': [total_responses, positive_count, negative_count,
                     f'{(positive_count/total_responses)*100:.1f}%',
                     f'{(negative_count/total_responses)*100:.1f}%']
        }
        summary_df = pd.DataFrame(summary_data)
        
        # Category distribution
        category_dist = all_results_df['التصنيف'].value_counts().reset_index()
        category_dist.columns = ['التصنيف', 'العدد']
        category_dist['النسبة'] = (category_dist['العدد'] / total_responses * 100).round(1).astype(str) + '%'

        # Write sheets
        summary_df.to_excel(writer, sheet_name='ملخص التحليل', index=False, startrow=1)
        category_dist.to_excel(writer, sheet_name='ملخص التحليل', index=False, startrow=7)
        all_results_df.to_excel(writer, sheet_name='جميع النتائج', index=False)
        
        # Separate positive and negative experiences
        positive_df = all_results_df[all_results_df['النوع'] == 'إيجابي']
        negative_df = all_results_df[all_results_df['النوع'] == 'سلبي']
        positive_df.to_excel(writer, sheet_name='التجارب الإيجابية', index=False)
        negative_df.to_excel(writer, sheet_name='التجارب السلبية', index=False)

        # Create pivot tables
        pivot_sheet_name = 'تحليل التصنيفات'
        category_pivot = pd.pivot_table(
            all_results_df,
            values='الاستجابة',
            index=['التصنيف', 'التصنيف الفرعي'],
            columns=['النوع'],
            aggfunc='count',
            fill_value=0
        ).reset_index()
        category_pivot.to_excel(writer, sheet_name=pivot_sheet_name)

        # Get workbook and sheets
        workbook = writer.book
        
        # Add charts
        from openpyxl.chart import PieChart, BarChart, Reference
        
        # Summary sheet formatting
        summary_sheet = writer.sheets['ملخص التحليل']
        summary_sheet.sheet_view.rightToLeft = True
        
        # Add title
        summary_sheet['A1'] = 'ملخص تحليل تجارب الطلاب'
        summary_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
        summary_sheet.merge_cells('A1:B1')
        
        # Create pie chart for positive/negative distribution
        pie = PieChart()
        pie.title = 'توزيع التجارب'
        labels = Reference(summary_sheet, min_col=1, min_row=3, max_row=4)
        data = Reference(summary_sheet, min_col=2, min_row=3, max_row=4)
        pie.add_data(data)
        pie.set_categories(labels)
        summary_sheet.add_chart(pie, "D2")

        # Create bar chart for category distribution
        bar = BarChart()
        bar.title = 'توزيع التصنيفات'
        bar.type = "col"
        bar.style = 10
        bar.y_axis.title = 'العدد'
        bar.x_axis.title = 'التصنيف'
        
        data = Reference(summary_sheet, min_col=2, min_row=8, max_row=8+len(category_dist))
        cats = Reference(summary_sheet, min_col=1, min_row=8, max_row=8+len(category_dist))
        bar.add_data(data)
        bar.set_categories(cats)
        summary_sheet.add_chart(bar, "D15")

        # Apply conditional formatting and styling to all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            worksheet.sheet_view.rightToLeft = True
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        # Skip merged cells
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                            continue
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # Only adjust if we found a valid column letter
                if column and not isinstance(column[0], openpyxl.cell.cell.MergedCell):
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Add header styling
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True, size=12)
                cell.fill = openpyxl.styles.PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='medium'))
            
            # Add zebra striping
            for row in range(2, worksheet.max_row + 1):
                if row % 2 == 0:
                    for cell in worksheet[row]:
                        cell.fill = openpyxl.styles.PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            
            # Add conditional formatting for positive/negative
            if 'النوع' in [cell.value for cell in worksheet[1]]:
                type_col = None
                for idx, cell in enumerate(worksheet[1], 1):
                    if cell.value == 'النوع':
                        type_col = idx
                        break
                
                if type_col:
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=type_col)
                        if cell.value == 'إيجابي':
                            cell.font = openpyxl.styles.Font(color='4CAF50')
                        elif cell.value == 'سلبي':
                            cell.font = openpyxl.styles.Font(color='F44336')

    # Reset pointer and get value
    output.seek(0)
    excel_data = output.getvalue()
    
    # Calculate top categories and their percentages
    category_stats = {}
    for r in st.session_state.results:
        category = r.get('classification', {}).get('category', '')
        type_ = r.get('classification', {}).get('type', '')
        if category:
            if category not in category_stats:
                category_stats[category] = {'total': 0, 'positive': 0}
            category_stats[category]['total'] += 1
            if type_ == 'إيجابي':
                category_stats[category]['positive'] += 1
    
    # Calculate percentages and sort by total count
    for cat in category_stats:
        total = category_stats[cat]['total']
        positive = category_stats[cat]['positive']
        category_stats[cat]['positive_pct'] = (positive / total) * 100
        category_stats[cat]['negative_pct'] = ((total - positive) / total) * 100
    
    # Get top 4 categories by total count
    top_categories = sorted(category_stats.items(), key=lambda x: x[1]['total'], reverse=True)[:4]
    
    # Display enhanced summary section
    summary_html = f"""
        <div class="summary-container">
            <div class="summary-header">ملخص التصنيف</div>
            <div class="summary-stats">
                <div class="stat-card positive-stat">
                    <div class="stat-number">{len(positive_experiences)}</div>
                    <div class="stat-label">تجارب إيجابية</div>
                </div>
                <div class="stat-card negative-stat">
                    <div class="stat-number">{len(negative_experiences)}</div>
                    <div class="stat-label">تجارب سلبية</div>
                </div>
            </div>
            <div class="top-categories">
                <div class="top-categories-header">أعلى 4 تصنيفات</div>
                <div class="category-grid">"""
    
    # Add category cards with percentages
    for cat, stats in top_categories:
        summary_html += f"""
            <div class="category-card">
                <div class="category-count">{stats['total']}</div>
                <div class="category-name">{cat}</div>
                <div class="category-percentages">
                    <span class="positive-pct">{stats['positive_pct']:.1f}% إيجابي</span>
                    <span class="negative-pct">{stats['negative_pct']:.1f}% سلبي</span>
                </div>
            </div>"""
    
    summary_html += """
                </div>
            </div>
        </div>
    """
    
    # Add CSS for percentages
    st.markdown("""
        <style>
        .category-percentages {
            margin-top: 5px;
            font-size: 0.8em;
            display: flex;
            flex-direction: column;
            gap: 2px;
            text-align: center;
        }
        
        .positive-pct {
            color: #4CAF50;
        }
        
        .negative-pct {
            color: #f44336;
        }
        
        .category-card {
            padding: 15px 10px !important;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown(summary_html, unsafe_allow_html=True)
    
    # Add download button
    st.download_button(
        "تحميل النتائج كملف Excel",
        excel_data,
        "classification_results.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key='download_excel',
        use_container_width=True,
    )
    
    # Add page switching button
    if st.button("عرض التفاصيل", use_container_width=True):
        switch_page("detailed_results")
